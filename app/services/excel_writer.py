from openpyxl import Workbook, load_workbook
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)

COMPANY_FILE = DATA_DIR / "companies.xlsx"


def _get_or_create_wb(path: Path):
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        # remove default sheet
        wb.remove(wb.active)
    return wb


def _get_or_create_sheet(wb, sheet_name: str, headers: list[str]):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
    return ws


def write_company_with_directors_and_meeting(company):
    wb = _get_or_create_wb(COMPANY_FILE)
    # ---------- Companies sheet ----------
    companies_ws = _get_or_create_sheet(
        wb,
        "Companies",
        [
            "id",
            "company_name",
            "company_type",
            "incorporation_date",
            "registered_office_address",
        ],
    )

    company_id = companies_ws.max_row  # simple incremental ID

    companies_ws.append(
        [
            company_id,
            company.companyName,
            company.companyType,
            company.incorporationDate.isoformat(),
            company.registeredOffice,
        ]
    )

    # ---------- Directors sheet ----------
    directors_ws = _get_or_create_sheet(
        wb,
        "Directors",
        [
            "id",
            "company_id",
            "name",
            "din",
            "designation",
            "address",
        ],
    )

    for director in company.directors:
        director_id = directors_ws.max_row
        directors_ws.append(
            [
                director_id,
                company_id,
                director.name,
                director.din,
                director.designation,
                director.address,
            ]
        )

    # ---------- Meetings sheet ----------
    meetings_ws = _get_or_create_sheet(
        wb,
        "Meetings",
        [
            "id",
            "company_id",
            "meeting_type",
            "attendance_mode",
            "serial_number",
            "financial_year",
            "signatory",
            "meeting_link",
            "meeting_id",
            "meeting_passcode",
            "day",
            "date",
            "time",
            "venue",
            "shorter_notice",
        ],
    )

    meeting = company.meeting
    meeting_id = meetings_ws.max_row

    meetings_ws.append(
        [
            meeting_id,
            company_id,
            meeting.meetingType,
            meeting.attendanceMode,
            meeting.serialNumber,
            meeting.financialYear,
            meeting.signatory,
            str(meeting.meetingLink),
            meeting.meetingId,
            meeting.meetingPasscode,
            meeting.meetingDay,
            meeting.meetingDate.isoformat(),
            meeting.meetingTime.strftime("%H:%M"),
            meeting.venue,
            meeting.shorterNotice,
        ]
    )

    wb.save(COMPANY_FILE)

    return company_id
