from docxtpl import DocxTemplate
import openpyxl
from pathlib import Path
import copy
from docx.oxml.ns import qn


def safe_filename(name: str) -> str:
    return "".join(c if c.isalnum() or c in " _-" else "_" for c in name)


def load_key_value_sheet(workbook, sheet_name="Meeting"):
    sheet = workbook[sheet_name]
    return {
        row[0]: row[1]
        for row in sheet.iter_rows(values_only=True)
        if row[0]
    }


def load_directors_from_excel(workbook, sheet_name="Directors"):
    sheet = workbook[sheet_name]

    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]

    directors = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))

        if not record.get("NAME"):
            continue

        directors.append({
            "name": record["NAME"],
            "din": str(record["DIN"]),
            "designation": record["DESIGNATION"],
            "address": record["ADDRESS"],
        })

    return directors


def load_agenda_from_excel(workbook, sheet_name="Agenda"):
    sheet = workbook[sheet_name]

    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]

    agenda_items = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        agenda_items.append(dict(zip(headers, row)))

    return agenda_items



def generate_snc_documents(
    template_path,
    output_dir,
    base_context,
    directors,
    company_name_safe,
    safe_filename
):
    """
    Generates one Shorter Notice Consent document per director
    """

    for director in directors:
        snc_doc = DocxTemplate(template_path)

        snc_context = base_context.copy()
        snc_context.update({
            "DIRECTOR_NAME": director["name"],
            "DIRECTOR_DIN": director["din"],
            "DIRECTOR_DESIGNATION": director["designation"],
            "DIRECTOR_ADDRESS": director["address"],
        })

        director_name_safe = safe_filename(director["name"])
        filename = f"SNC_{company_name_safe}_{director_name_safe}.docx"

        snc_doc.render(snc_context)
        snc_doc.save(output_dir / filename)



def load_resolution_index(path: Path, meeting_type: str):
    wb = openpyxl.load_workbook(path)
    sheet = wb["Index"]

    index = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        resolution_id, resolution_title, resolution_file = row

         # 🔐 Enforce .docx extension
        if not resolution_file.lower().endswith(".docx"):
            resolution_file += ".docx"

        index[resolution_id] = {
            "title": resolution_title,
            "template": resolution_file
        }

    return index
 
def build_resolution_subdoc(parent_doc, resolution_template_path, context):
    """
    Safely builds a resolution subdoc preserving formatting
    """
    resolution_subdoc = parent_doc.new_subdoc()

    tpl = DocxTemplate(resolution_template_path)
    tpl.render(context)

    for element in tpl.docx._body._element:
        if element.tag == qn("w:sectPr"):
            continue
        resolution_subdoc._body._element.append(copy.deepcopy(element))

    return resolution_subdoc
